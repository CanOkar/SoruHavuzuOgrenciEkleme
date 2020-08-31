from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
import openpyxl
import getpass

print("- Zafer Sınav sistemine giriş için lütfen tc kimlik ve parolanızı girin")
uname = input("- Tc Kimlik: ")
passwrd = getpass.getpass("- Parola: ")
print("- Öğrenci kayıtlarını içeren excel doyanızın ismini ve yolunu girin. Örnek: C:\ogrenciler.xlsx")
print("- Eğer program dosyanızla excel dosyanız aynı dizinde ise sadece excel dosyanızın adını girin. Örnek: ogrenciler.xlsx")
excelFile = input("- Excel Dosyanız: ")
###########################Login Part###########################
driver = webdriver.Chrome('/snap/bin/chromium.chromedriver')
driver.get("http://sinav.zaferkoleji.com.tr")

enterUsername = driver.find_element_by_xpath('//*[@id="app"]/div[2]/div[2]/div/div/div/div[2]/div[1]/input')
enterUsername.send_keys(uname)
enterPassword = driver.find_element_by_xpath('//*[@id="app"]/div[2]/div[2]/div/div/div/div[2]/div[2]/input')
enterPassword.send_keys(passwrd)
Loginbutton = driver.find_element_by_xpath('//*[@id="app"]/div[2]/div[2]/div/div/div/div[2]/div[3]/button[1]')
Loginbutton.click()
time.sleep(2)
###########################Login Part###########################

# read from excel file
wb = openpyxl.load_workbook(excelFile)
sheet = wb['Sheet1']
rowCount = sheet.max_row
print("- Toplam " , rowCount , " satırlık excel dosyasındaki kayıtların sisteme girişi yapılacak")
for i in range(2, rowCount + 1):
    driver.get("http://sinav.zaferkoleji.com.tr/students")
    time.sleep(1)

    for x in range(1, 8):
        if x == 1:
            sn = sheet.cell(row=i, column=x).value
            schoolnumber = driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/div[1]/div/div/div/div/div[2]/form/div[1]/div[3]/div[1]/div/input')
            schoolnumber.send_keys(sn)
        elif x == 2:
            fn = sheet.cell(row=i, column=x).value
            name = driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/div[1]/div/div/div/div/div[2]/form/div[1]/div[1]/div[1]/div/input')
            name.send_keys(fn)
        elif x == 3:
            ln = sheet.cell(row=i, column=x).value
            lastname = driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/div[1]/div/div/div/div/div[2]/form/div[1]/div[1]/div[2]/div/input')
            lastname.send_keys(ln)
        elif x == 4:
            tck = sheet.cell(row=i, column=x).value
            tc = driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/div[1]/div/div/div/div/div[2]/form/div[1]/div[2]/div[2]/div/input')
            tc.send_keys(tck)
        elif x == 5:
            sch = sheet.cell(row=i, column=x).value
            selectSchool = driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/div[1]/div/div/div/div/div[2]/form/div[1]/div[3]/div[2]/div/div[1]')
            selectSchool.click()

            enterSchoolName = driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/div[1]/div/div/div/div/div[2]/form/div[1]/div[3]/div[2]/div/div[2]/input')
            enterSchoolName.send_keys(sch)
            enterSchoolName.send_keys(Keys.RETURN)
            time.sleep(1)
        elif x == 6:
            cls = sheet.cell(row=i, column=x).value
            selectClass = driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/div[1]/div/div/div/div/div[2]/form/div[1]/div[4]/div/div/div[1]')
            selectClass.click()

            enterClassName = driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/div[1]/div/div/div/div/div[2]/form/div[1]/div[4]/div/div/div[2]/input')
            enterClassName.send_keys(cls)
            enterClassName.send_keys(Keys.RETURN)
        elif x == 7:
            em = sheet.cell(row=i, column=x).value
            email = driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/div[1]/div/div/div/div/div[2]/form/div[1]/div[2]/div[1]/div/input')
            email.send_keys(em)
        time.sleep(1)
    saveButton = driver.find_element_by_xpath('//*[@id="app"]/div[2]/div/div/div/div[1]/div/div/div/div/div[2]/form/div[2]/div/button')
    saveButton.click()
    time.sleep(3)
